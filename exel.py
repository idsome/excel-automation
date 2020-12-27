from openpyxl import load_workbook
#direction to the two workbooks
read_file = "./templates/book1/book1.xlsx"
write_file ="./templates/book2/book2.xlsx"
#loading the location and selecting a sheet
Wb = load_workbook(filename=read_file)
sheet1 =Wb["Sheet1"]
wb2 = load_workbook(filename=write_file)
sheet2 =wb2['Sheet1']
#an outer for loop that scans the two selected rows in the first excel workbook
for raw in range(1 ,6):
    y = sheet1.cell(row=raw , column=1)
    c = sheet1.cell(row=raw , column=2)
    #an inner for loop that scans the two selected rows in the second workbook
    for items in range(1 ,6):
        z = sheet2.cell (row=items , column=1)
        l = sheet2.cell(row=items ,column=2)
        # an if statment to compare the first rows in the two workbooks
        if str(y.value) == str(z.value):
            #if the the firts rows matches this line copys the content of the second row from the first to the second workbook
            l.value =c.value
#this line saves the new workbook in the same location as this code 
wb2.save(filename='book2.xlsx')
print('done')

    